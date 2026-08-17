# ============================================================
# core/reporting/excel_builder.py
#
# Plain openpyxl styling helpers shared by both report types. Kept to
# simple, reusable primitives (colored P&L cells, currency/percentage
# number formats, auto column widths, freeze panes, autofilter) rather
# than a templating system -- these are point-in-time snapshots, not
# live templates, so cell values are colored once at build time
# instead of via openpyxl's dynamic ConditionalFormatting rules.
# ============================================================

import re
from io import BytesIO

import pandas as pd
import pytz
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

IST = pytz.timezone("Asia/Kolkata")

# Values are stored as plain 0-100 numbers (e.g. 62.3), NOT Excel's
# native 0-1 percent representation -- deliberately using a literal-
# suffix custom format instead of the built-in "0.0%" (which rescales
# by x100 and would double-convert an already-0-100 value).
CURRENCY_FORMAT = '"₹"#,##0.00'
PERCENT_FORMAT  = '0.0"%"'

_POS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_POS_FONT = Font(color="006100")
_NEG_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_NEG_FONT = Font(color="9C0006")


def slugify_strategy(strategy: str) -> str:
    """'RSI + MA' -> 'RSI-MA', for use in downloaded filenames."""
    return re.sub(r"[^a-zA-Z0-9]+", "-", strategy.strip()).strip("-")


def _excel_safe(val):
    """openpyxl rejects tz-aware datetimes outright -- convert to IST
    wall-clock and drop tzinfo (the report is IST-oriented anyway)."""
    if pd.isna(val):
        return None
    if isinstance(val, pd.Timestamp):
        return (val.tz_convert(IST) if val.tzinfo is not None else val).to_pydatetime().replace(tzinfo=None)
    if hasattr(val, "tzinfo") and val.tzinfo is not None:
        return val.astimezone(IST).replace(tzinfo=None)
    return val


def write_dataframe(
    ws, df: pd.DataFrame, start_row: int = 1,
    pnl_columns: tuple = (), currency_columns: tuple = (), percent_columns: tuple = (),
    freeze_col: int = None, autofilter: bool = False,
) -> int:
    """
    Writes a DataFrame to `ws` starting at `start_row` (1-indexed):
    bold header row + data rows, with green/red fill+font on
    `pnl_columns` (by header name, based on each cell's own sign),
    currency/percent number formats on the named columns, optional
    freeze_panes at column `freeze_col`, optional autofilter, and
    auto column widths (header/content length + 2, capped at 40).
    Returns the next free row.
    """
    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="(no data for this period)")
        return start_row + 2

    headers = list(df.columns)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=str(h)).font = Font(bold=True)

    for i, (_, record) in enumerate(df.iterrows(), start=start_row + 1):
        for j, h in enumerate(headers, start=1):
            val = _excel_safe(record[h])
            cell = ws.cell(row=i, column=j, value=val)
            if h in pnl_columns and isinstance(val, (int, float)):
                if val >= 0:
                    cell.fill, cell.font = _POS_FILL, _POS_FONT
                else:
                    cell.fill, cell.font = _NEG_FILL, _NEG_FONT
            if h in currency_columns and isinstance(val, (int, float)):
                cell.number_format = CURRENCY_FORMAT
            if h in percent_columns and isinstance(val, (int, float)):
                cell.number_format = PERCENT_FORMAT

    for j, h in enumerate(headers, start=1):
        sample = df[h].astype(str).values[:2000]
        max_len = max([len(str(h))] + [len(v) for v in sample])
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 2, 40)

    end_row = start_row + len(df)
    if freeze_col:
        ws.freeze_panes = ws.cell(row=start_row + 1, column=freeze_col + 1).coordinate
    if autofilter:
        ws.auto_filter.ref = f"{get_column_letter(1)}{start_row}:{get_column_letter(len(headers))}{end_row}"

    return end_row + 2


def write_overview_block(ws, meta: dict, kpis: dict, start_row: int = 1) -> int:
    """Key/value block for the Overview sheet: metadata rows, a blank
    line, then KPI rows (P&L-looking keys get currency format + color)."""
    row = start_row
    for k, v in meta.items():
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
        row += 1
    row += 1
    for k, v in kpis.items():
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=v)
        if isinstance(v, (int, float)) and "p&l" in k.lower():
            cell.number_format = CURRENCY_FORMAT
            if v >= 0:
                cell.fill, cell.font = _POS_FILL, _POS_FONT
            else:
                cell.fill, cell.font = _NEG_FILL, _NEG_FONT
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    return row + 1


def add_equity_curve_chart(ws, data_start_row: int, data_end_row: int, date_col: int, value_col: int, anchor_cell: str = "E2"):
    """Embeds a LineChart plotting the cumulative-P&L column over the
    date column. data_start_row/data_end_row are the DATA rows (header
    is assumed to sit one row above data_start_row)."""
    if data_end_row < data_start_row:
        return
    chart = LineChart()
    chart.title = "Cumulative Net P&L"
    chart.style = 2
    chart.y_axis.title = "₹"
    chart.x_axis.title = "Date"
    data = Reference(ws, min_col=value_col, min_row=data_start_row - 1, max_row=data_end_row)
    cats = Reference(ws, min_col=date_col, min_row=data_start_row, max_row=data_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 18, 9
    ws.add_chart(chart, anchor_cell)


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
