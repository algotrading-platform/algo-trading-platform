# ============================================================
# core/reporting/signal_history_report.py
#
# Builds the 5-sheet Signal History Excel report end-to-end and
# returns the workbook as bytes, ready for st.download_button.
# Sheets: Overview, Strategy Summary, Symbol Frequency, Signal
# Conversion, Signal Log.
# ============================================================

from datetime import timedelta as _td, datetime

import pytz
from openpyxl import Workbook
from openpyxl.styles import Font

from core.database.db import get_paper_trades_for_report
from core.reporting.excel_builder import slugify_strategy, workbook_to_bytes, write_dataframe, write_overview_block
from core.reporting.periods import count_trading_days, get_period_bounds
from core.reporting.report_data import build_signal_history_dataset, compute_signal_conversion

try:
    from core.scheduler.signal_scheduler import NSE_HOLIDAYS
except Exception:
    NSE_HOLIDAYS = set()

IST = pytz.timezone("Asia/Kolkata")
MAX_ROWS = 200_000  # same defensive guardrail as the paper trading report


def build_signal_history_report(
    period: str, anchor=None, custom_start=None, custom_end=None,
    strategy: str = None, timeframe: str = None,
) -> tuple:
    """Returns (file_bytes, filename, stats_caption)."""
    start_utc, end_utc, label = get_period_bounds(period, anchor, custom_start, custom_end)
    data = build_signal_history_dataset(start_utc, end_utc, strategy, timeframe)
    signals = data["signals"]

    if len(signals) > MAX_ROWS:
        raise ValueError(
            f"This period has {len(signals):,} signals, above the {MAX_ROWS:,}-row "
            f"report limit. Narrow the date range or add a strategy filter."
        )

    trades     = get_paper_trades_for_report(start_utc, end_utc, strategy, timeframe)
    conversion = compute_signal_conversion(signals, trades)

    total_signals = len(signals)
    buy_count  = int((signals["Signal"] == "BUY").sum()) if total_signals else 0
    sell_count = int((signals["Signal"] == "SELL").sum()) if total_signals else 0
    symbols    = int(signals["Stock"].nunique()) if total_signals else 0

    start_ist = start_utc.astimezone(IST).date()
    end_ist_incl = end_utc.astimezone(IST).date() - _td(days=1)

    wb = Workbook()

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = "Overview"
    meta = {
        "Report Type":           "Signal History",
        "Period":                period.capitalize(),
        "Date Range (IST)":      f"{start_ist} to {end_ist_incl}",
        "Strategy Filter":       strategy or "All Strategies",
        "Timeframe Filter":      timeframe or "All Timeframes",
        "Trading Days in Period": count_trading_days(start_utc, end_utc, NSE_HOLIDAYS),
        "Generated At (IST)":    datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
    }
    kpis = {
        "Total Signals":   total_signals,
        "BUY Signals":     buy_count,
        "SELL Signals":    sell_count,
        "Distinct Symbols": symbols,
    }
    write_overview_block(ws, meta, kpis, start_row=1)

    # ── Sheet 2: Strategy Summary ──
    ws2 = wb.create_sheet("Strategy Summary")
    write_dataframe(ws2, data["by_strategy"], freeze_col=2, autofilter=True)

    # ── Sheet 3: Symbol Frequency ──
    ws3 = wb.create_sheet("Symbol Frequency")
    write_dataframe(ws3, data["by_symbol"], freeze_col=1, autofilter=True)

    # ── Sheet 4: Signal Conversion (headline for this report — is
    # signal generation actually turning into trades, independent of
    # whether RMS caps let it) ──
    ws4 = wb.create_sheet("Signal Conversion")
    write_dataframe(ws4, conversion, percent_columns=("conversion_rate",), freeze_col=2, autofilter=True)

    # ── Sheet 5: Signal Log (full detail) ──
    ws5 = wb.create_sheet("Signal Log")
    write_dataframe(ws5, signals, freeze_col=2, autofilter=True)

    file_bytes = workbook_to_bytes(wb)
    strategy_slug = f"_{slugify_strategy(strategy)}" if strategy else ""
    filename = f"signal_history_report_{label}{strategy_slug}.xlsx"
    stats_caption = f"{total_signals} signals ({buy_count} BUY / {sell_count} SELL), {symbols} symbols"
    return file_bytes, filename, stats_caption
