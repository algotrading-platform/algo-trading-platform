# ============================================================
# core/reporting/paper_trading_report.py
#
# Builds the 7-sheet Paper Trading Excel report end-to-end and
# returns the workbook as bytes, ready for st.download_button.
# Sheets: Overview, Strategy Performance, Equity Curve (+chart),
# Symbol Breakdown, Day & Time Patterns, Signal Conversion, Trade Log.
# ============================================================

from datetime import datetime

import pytz
from openpyxl import Workbook
from openpyxl.styles import Font

from core.database.db import get_signals_for_report
from core.reporting.excel_builder import (
    add_equity_curve_chart, slugify_strategy, workbook_to_bytes,
    write_dataframe, write_overview_block,
)
from core.reporting.periods import count_trading_days, get_period_bounds
from core.reporting.report_data import (
    build_paper_trading_dataset, build_strategy_performance_table, compute_signal_conversion,
)

try:
    from core.scheduler.signal_scheduler import NSE_HOLIDAYS
except Exception:
    NSE_HOLIDAYS = set()

IST = pytz.timezone("Asia/Kolkata")

# Defensive guardrail only -- not expected to trigger at this app's
# real trade volume (see the reporting design's scaling assessment).
# Shows a friendly warning instead of ever silently building a huge
# workbook.
MAX_ROWS = 200_000


def build_paper_trading_report(
    period: str, anchor=None, custom_start=None, custom_end=None,
    strategy: str = None, timeframe: str = None,
) -> tuple:
    """Returns (file_bytes, filename, stats_caption)."""
    start_utc, end_utc, label = get_period_bounds(period, anchor, custom_start, custom_end)
    data = build_paper_trading_dataset(start_utc, end_utc, strategy, timeframe)
    trades = data["trades"]

    if len(trades) > MAX_ROWS:
        raise ValueError(
            f"This period has {len(trades):,} closed trades, above the {MAX_ROWS:,}-row "
            f"report limit. Narrow the date range or add a strategy filter."
        )

    signals    = get_signals_for_report(start_utc, end_utc, strategy, timeframe)
    conversion = compute_signal_conversion(signals, trades)

    total_trades  = len(trades)
    total_net     = round(float(trades["net_pnl"].sum()), 2) if total_trades else 0.0
    total_gross   = round(float(trades["pnl"].sum()), 2) if total_trades else 0.0
    total_charges = round(float(trades["charges"].sum()), 2) if total_trades else 0.0
    wins          = int((trades["pnl"] > 0).sum()) if total_trades else 0
    win_rate      = round(wins / total_trades * 100, 1) if total_trades else 0.0
    charges_drag  = round(total_charges / abs(total_gross) * 100, 1) if total_gross else 0.0

    start_ist = start_utc.astimezone(IST).date()
    end_ist_incl = end_utc.astimezone(IST).date()
    from datetime import timedelta as _td
    end_ist_incl = end_ist_incl - _td(days=1)

    wb = Workbook()

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = "Overview"
    meta = {
        "Report Type":          "Paper Trading",
        "Period":                period.capitalize(),
        "Date Range (IST)":      f"{start_ist} to {end_ist_incl}",
        "Strategy Filter":       strategy or "All Strategies",
        "Timeframe Filter":      timeframe or "All Timeframes",
        "Trading Days in Period": count_trading_days(start_utc, end_utc, NSE_HOLIDAYS),
        "Generated At (IST)":    datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
    }
    kpis = {
        "Total Trades":     total_trades,
        "Win Rate (%)":     win_rate,
        "Gross P&L":        total_gross,
        "Net P&L":          total_net,
        "Total Charges":    total_charges,
        "Charges Drag (%)": charges_drag,
    }
    next_row = write_overview_block(ws, meta, kpis, start_row=1)

    by_strategy = data["by_strategy"]
    if not by_strategy.empty:
        bs = by_strategy.copy()
        bs["win_rate"] = (bs["wins"] / bs["trades"] * 100).round(1)
        bs = bs[["strategy", "trades", "wins", "losses", "win_rate", "total_pnl", "total_net_pnl", "total_charges"]]
        ws.cell(row=next_row, column=1, value="Per-Strategy Breakdown").font = Font(bold=True, size=12)
        next_row += 1
        write_dataframe(
            ws, bs, start_row=next_row,
            pnl_columns=("total_pnl", "total_net_pnl"),
            currency_columns=("total_pnl", "total_net_pnl", "total_charges"),
            percent_columns=("win_rate",),
        )

    # ── Sheet 2: Strategy Performance ──
    ws2 = wb.create_sheet("Strategy Performance")
    perf = build_strategy_performance_table(trades, by_strategy)
    write_dataframe(
        ws2, perf,
        pnl_columns=("total_pnl", "total_net_pnl", "avg_win", "avg_loss", "best_trade", "worst_trade"),
        currency_columns=("total_pnl", "total_net_pnl", "total_charges", "avg_win", "avg_loss", "best_trade", "worst_trade"),
        percent_columns=("win_rate",), freeze_col=1, autofilter=True,
    )

    # ── Sheet 3: Equity Curve (+ chart) ──
    ws3 = wb.create_sheet("Equity Curve")
    eq = data["equity_curve"]
    write_dataframe(ws3, eq, currency_columns=("daily_net_pnl", "cumulative_net_pnl"),
                    pnl_columns=("daily_net_pnl", "cumulative_net_pnl"))
    if not eq.empty:
        add_equity_curve_chart(ws3, data_start_row=2, data_end_row=1 + len(eq), date_col=1, value_col=3, anchor_cell="E2")

    # ── Sheet 4: Symbol Breakdown ──
    ws4 = wb.create_sheet("Symbol Breakdown")
    write_dataframe(
        ws4, data["by_symbol"],
        pnl_columns=("net_pnl", "gross_pnl", "avg_pnl", "best_trade", "worst_trade"),
        currency_columns=("net_pnl", "gross_pnl", "avg_pnl", "best_trade", "worst_trade"),
        percent_columns=("win_rate",), freeze_col=1, autofilter=True,
    )

    # ── Sheet 5: Day & Time Patterns ──
    ws5 = wb.create_sheet("Day & Time Patterns")
    r = 1
    ws5.cell(row=r, column=1, value="By Day of Week").font = Font(bold=True, size=12)
    r += 1
    r = write_dataframe(ws5, data["by_day"], start_row=r, pnl_columns=("net_pnl",),
                        currency_columns=("net_pnl",), percent_columns=("win_rate",))
    ws5.cell(row=r, column=1, value="By Entry Hour (IST)").font = Font(bold=True, size=12)
    r += 1
    write_dataframe(ws5, data["by_hour"], start_row=r, pnl_columns=("net_pnl",),
                    currency_columns=("net_pnl",), percent_columns=("win_rate",))

    # ── Sheet 6: Signal Conversion ──
    ws6 = wb.create_sheet("Signal Conversion")
    write_dataframe(ws6, conversion, percent_columns=("conversion_rate",), freeze_col=2, autofilter=True)

    # ── Sheet 7: Trade Log (full detail) ──
    ws7 = wb.create_sheet("Trade Log")
    write_dataframe(
        ws7, trades, pnl_columns=("pnl", "net_pnl"),
        currency_columns=("entry_price", "exit_price", "stop_loss", "target", "pnl", "net_pnl", "charges", "risk_amount"),
        freeze_col=4, autofilter=True,
    )

    file_bytes = workbook_to_bytes(wb)
    strategy_slug = f"_{slugify_strategy(strategy)}" if strategy else ""
    filename = f"paper_trading_report_{label}{strategy_slug}.xlsx"
    stats_caption = f"{total_trades} trades, {win_rate}% win rate, net P&L ₹{total_net:,.2f}"
    return file_bytes, filename, stats_caption
